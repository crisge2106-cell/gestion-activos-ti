#!/usr/bin/env python3
"""
Script para generar Excel de solicitud de compra usando plantilla
Uso: python3 generar_solicitud_excel.py <id_solicitud> <output_path>
"""

import sys
import json
from openpyxl import load_workbook
from openpyxl.styles import Font
from datetime import datetime

def generar_excel_solicitud(sol_data, output_path, plantilla_path='plantillas/ADM-FO-01_Requerimiento.xlsx'):
    """
    Genera un Excel de solicitud rellenando la plantilla con datos

    sol_data: dict con {numero, fecha, usuario, items: []}
    """
    try:
        # Cargar plantilla
        wb = load_workbook(plantilla_path)
        ws = wb.active

        # Rellenar datos generales
        # Fila 6: Solicitado por
        ws['D6'] = sol_data.get('usuario', '').upper()
        ws['D6'].font = Font(bold=True)

        # Fila 8: Área
        ws['D8'] = 'TECNOLOGIAS DE LA INFORMACIÓN'
        ws['D8'].font = Font(bold=True)

        # Requerimiento Nº (Fila 6, columna I-J)
        ws['J6'] = sol_data.get('numero', '')

        # Fecha de solicitud (Fila 8, columna J)
        fecha = sol_data.get('fecha', '')
        if fecha:
            # Convertir ISO a formato más legible
            try:
                fecha_obj = datetime.fromisoformat(fecha.replace('Z', '+00:00'))
                fecha_str = fecha_obj.strftime('%Y-%m-%d')
            except:
                fecha_str = fecha[:10]
        else:
            fecha_str = datetime.now().strftime('%Y-%m-%d')

        ws['J8'] = fecha_str
        ws['J8'].font = Font(bold=True)

        # Items - Encabezados en fila 10
        # Fila 11 en adelante: items
        row = 11
        for item_idx, item in enumerate(sol_data.get('items', []), 1):
            # ITEM
            ws[f'B{row}'] = item_idx

            # CANTIDAD
            ws[f'C{row}'] = item.get('cantidad', 1)

            # UNIDAD
            ws[f'D{row}'] = 'UNIDAD'

            # DESCRIPCIÓN
            desc = item.get('descripcion', '')
            ws[f'E{row}'] = desc
            ws[f'E{row}'].font = Font(bold=True)

            # FECHA DE ENTREGA (Columna J)
            ws[f'J{row}'] = fecha_str

            # OBSERVACIÓN (Columna K)
            usuario_dest = item.get('usuarioDestino', '')
            if usuario_dest:
                ws[f'K{row}'] = f"Para: {usuario_dest}"

            row += 1

        # Guardar archivo
        wb.save(output_path)
        print(json.dumps({'success': True, 'path': output_path}))

    except Exception as e:
        print(json.dumps({'success': False, 'error': str(e)}), file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    if len(sys.argv) < 3:
        print(json.dumps({'success': False, 'error': 'Uso: python3 script.py <json_data> <output_path>'}), file=sys.stderr)
        sys.exit(1)

    try:
        sol_data = json.loads(sys.argv[1])
        output_path = sys.argv[2]
        generar_excel_solicitud(sol_data, output_path)
    except json.JSONDecodeError as e:
        print(json.dumps({'success': False, 'error': f'JSON inválido: {str(e)}'}), file=sys.stderr)
        sys.exit(1)
