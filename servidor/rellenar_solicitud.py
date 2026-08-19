#!/usr/bin/env python3
"""
Rellena la plantilla Excel con datos de la solicitud
Uso: python3 rellenar_solicitud.py <plantilla> <json_solicitud> <output>
"""

import sys
import json
from openpyxl import load_workbook
from datetime import datetime

def rellenar_plantilla(plantilla_path, solicitud_json, output_path):
    """Rellena la plantilla con los datos de la solicitud"""
    try:
        # Cargar datos
        solicitud = json.loads(solicitud_json)

        # Cargar plantilla
        wb = load_workbook(plantilla_path)
        ws = wb.active

        # Rellenar datos generales
        ws['D6'] = solicitud['usuario'].upper()  # Solicitado por
        ws['J6'] = solicitud['numero']  # Requerimiento Nº

        fecha = solicitud['fecha'][:10]  # Convertir ISO a YYYY-MM-DD
        ws['J8'] = fecha  # Fecha de solicitud

        # Rellenar items (comenzar en fila 11)
        row = 11
        for item in solicitud['items']:
            # Número de item (autoincremental)
            ws[f'B{row}'] = len(solicitud.get('items', []))  # Será reemplazado por el loop

            # Cantidad
            ws[f'C{row}'] = item['cantidad']

            # Unidad (siempre UNIDAD)
            ws[f'D{row}'] = 'UNIDAD'

            # Descripción (columnas E)
            ws[f'E{row}'] = item['descripcion']

            # Fecha de entrega
            ws[f'J{row}'] = fecha

            # Observación
            if item.get('usuarioDestino'):
                ws[f'K{row}'] = f"Para: {item['usuarioDestino']}"

            row += 1

        # Corregir números de item
        items = solicitud.get('items', [])
        for idx, item in enumerate(items):
            ws[f'B{11 + idx}'] = idx + 1

        # Guardar archivo
        wb.save(output_path)

        # Retornar éxito
        print(json.dumps({'success': True, 'path': output_path}))

    except Exception as e:
        print(json.dumps({'success': False, 'error': str(e)}), file=sys.stderr)
        sys.exit(1)

if __name__ == '__main__':
    if len(sys.argv) != 4:
        print(json.dumps({'success': False, 'error': 'Uso: rellenar_solicitud.py <plantilla> <json> <output>'}), file=sys.stderr)
        sys.exit(1)

    rellenar_plantilla(sys.argv[1], sys.argv[2], sys.argv[3])
